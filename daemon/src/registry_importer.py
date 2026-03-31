"""
Device Registry Importer
Bulk import devices from JSON, CSV, or XLSX sources.
"""

import json
import csv
import logging
from pathlib import Path
from typing import List, Dict, Any
from dataclasses import asdict

logger = logging.getLogger(__name__)

class RegistryImporter:
    """Import devices from various formats into DeviceRegistry."""

    def __init__(self, registry):
        self.registry = registry

    def import_json(self, file_path: Path) -> Dict[str, Any]:
        """
        Import devices from JSON file.
        Expected format:
        {
            "devices": [
                {
                    "device_id": "DEV001",
                    "name": "Device 1",
                    "hardware_class": "V4",
                    "ip_address": "192.168.1.10",
                    "current_version": "0.0.15V4"
                }
            ]
        }
        """
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)

            devices = data.get('devices', [])
            added = 0
            errors = []

            for device_data in devices:
                try:
                    self.registry.add_device(
                        device_id=device_data['device_id'],
                        name=device_data.get('name', device_data['device_id']),
                        hardware_class=device_data['hardware_class'],
                        ip_address=device_data['ip_address'],
                        current_version=device_data.get('current_version', 'unknown')
                    )
                    added += 1
                except Exception as e:
                    errors.append(f"Device {device_data.get('device_id', 'unknown')}: {str(e)}")

            return {
                "ok": True,
                "added": added,
                "total": len(devices),
                "errors": errors if errors else None
            }
        except Exception as e:
            logger.error(f"Failed to import JSON: {e}")
            return {"ok": False, "error": str(e)}

    def import_csv(self, file_path: Path) -> Dict[str, Any]:
        """
        Import devices from CSV file.
        Expected columns: device_id, name, hardware_class, ip_address, current_version
        """
        try:
            added = 0
            errors = []

            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return {"ok": False, "error": "Empty CSV file"}

                required_fields = {'device_id', 'hardware_class', 'ip_address'}
                if not required_fields.issubset(set(reader.fieldnames)):
                    return {
                        "ok": False,
                        "error": f"Missing required columns: {required_fields}"
                    }

                for row in reader:
                    try:
                        self.registry.add_device(
                            device_id=row['device_id'].strip(),
                            name=row.get('name', row['device_id']).strip(),
                            hardware_class=row['hardware_class'].strip(),
                            ip_address=row['ip_address'].strip(),
                            current_version=row.get('current_version', 'unknown').strip()
                        )
                        added += 1
                    except Exception as e:
                        errors.append(f"Row {row.get('device_id', 'unknown')}: {str(e)}")

            return {
                "ok": True,
                "added": added,
                "errors": errors if errors else None
            }
        except Exception as e:
            logger.error(f"Failed to import CSV: {e}")
            return {"ok": False, "error": str(e)}

    def import_xlsx(self, file_path: Path) -> Dict[str, Any]:
        """
        Import devices from XLSX file.
        Expected columns: device_id, name, hardware_class, ip_address, current_version
        Requires openpyxl library.
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "ok": False,
                "error": "openpyxl not installed. Install with: pip install openpyxl"
            }

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            added = 0
            errors = []

            # Read header
            headers = [cell.value for cell in ws[1]]
            required_fields = {'device_id', 'hardware_class', 'ip_address'}

            if not required_fields.issubset(set(h.lower() for h in headers if h)):
                return {
                    "ok": False,
                    "error": f"Missing required columns: {required_fields}"
                }

            # Map columns
            col_map = {h.lower(): i for i, h in enumerate(headers) if h}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    device_id = row[col_map.get('device_id', 0)]
                    if not device_id:
                        continue  # Skip empty rows

                    self.registry.add_device(
                        device_id=str(device_id).strip(),
                        name=str(row[col_map.get('name', col_map.get('device_id', 0))] or device_id).strip(),
                        hardware_class=str(row[col_map.get('hardware_class', 0)]).strip(),
                        ip_address=str(row[col_map.get('ip_address', 0)]).strip(),
                        current_version=str(row[col_map.get('current_version', 0)] or 'unknown').strip()
                    )
                    added += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

            return {
                "ok": True,
                "added": added,
                "errors": errors if errors else None
            }
        except Exception as e:
            logger.error(f"Failed to import XLSX: {e}")
            return {"ok": False, "error": str(e)}
